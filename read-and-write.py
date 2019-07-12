import openpyxl
wb=openpyxl.load_workbook('example.xlsx')
ws=wb.active
print(wb.sheetnames)  #打印表单   不能用之前的wb.get_sheet_names
print(ws.title)     #打印该表单名字
print(ws['A1'].value)   #获取单元格的值
print(ws.cell(row=1,column=2).value)    #工作表的cell()方法并传递它row=1并column=2获取Cell单元格的对象B1

c=ws['B1']
print(c.coordinate)

print(ws[1])
print(ws['A'])

print(ws.max_column)    #知晓工作表的最大列
print(ws.max_row)       #知晓工作表的最大行

from openpyxl.utils import get_column_letter,column_index_from_string   #把列字母转化为数字//把列数字转化为字母
print(column_index_from_string('AA'))    #获得AA的列数字
print(get_column_letter(27))

print(ws['A1:B3'])                       #切片单元格   元组 ((A1 B1 C1),(A2 B2 C2),(A3 B3 C3))
########遍历单元格
for rowOfCellObjects in ws['A1':'C3']:
    print(rowOfCellObjects)        #从第一列到第n列  元组 ((A1 ，B1， C1),(A2， B2， C2),(A3， B3， C3))
    for cellObj in rowOfCellObjects:
        print(cellObj.coordinate, cellObj.value)
        print('--- END OF ROW ---')
